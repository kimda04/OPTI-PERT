from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.services.project_service import get_project


def generate_excel():
    project = get_project()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Proyecto PERT"

    header_fill = PatternFill(fill_type="solid", fgColor="6C63FF")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [
        "Actividad",
        "Descripción",
        "Optimista",
        "Probable",
        "Pesimista",
        "Tiempo Esperado",
        "Varianza",
        "ES",
        "EF",
        "LS",
        "LF",
        "Slack",
        "Ruta Crítica",
        "Predecesores"
    ]

    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column)
        cell.value = title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2

    for activity in project.activities:

        sheet.cell(row=row, column=1).value = activity.name
        sheet.cell(row=row, column=2).value = activity.description
        sheet.cell(row=row, column=3).value = activity.optimistic
        sheet.cell(row=row, column=4).value = activity.most_likely
        sheet.cell(row=row, column=5).value = activity.pessimistic
        sheet.cell(row=row, column=6).value = activity.expected_time
        sheet.cell(row=row, column=7).value = activity.variance
        sheet.cell(row=row, column=8).value = activity.es
        sheet.cell(row=row, column=9).value = activity.ef
        sheet.cell(row=row, column=10).value = activity.ls
        sheet.cell(row=row, column=11).value = activity.lf
        sheet.cell(row=row, column=12).value = activity.slack
        sheet.cell(row=row, column=13).value = (
            "Sí" if activity.name in project.critical_path else "No"
        )
        sheet.cell(row=row, column=14).value = ", ".join(activity.predecessors)

        row += 1

    row += 2

    sheet.cell(row=row, column=1).value = "Duración Total"
    sheet.cell(row=row, column=2).value = project.total_duration

    row += 1

    sheet.cell(row=row, column=1).value = "Ruta Crítica"
    sheet.cell(row=row, column=2).value = " → ".join(project.critical_path)

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 3

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer